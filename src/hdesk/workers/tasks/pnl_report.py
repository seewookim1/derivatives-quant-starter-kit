"""EOD P&L 보고서 생성 태스크 - openpyxl"""

from __future__ import annotations

import asyncio
import logging
from datetime import date
from pathlib import Path

logger = logging.getLogger(__name__)

REPORTS_DIR = Path("reports")


def generate_pnl_report() -> None:
    """EOD P&L Excel 보고서 생성 (동기 래퍼)."""
    REPORTS_DIR.mkdir(exist_ok=True)
    asyncio.run(_generate_async())


async def _generate_async() -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    from hdesk.data.database import get_db_session
    from hdesk.data.repositories.market_data_repo import MarketDataRepository
    from hdesk.data.repositories.position_repo import PositionRepository
    from hdesk.data.repositories.risk_repo import RiskRepository

    today_str = date.today().strftime("%Y%m%d")
    output_path = REPORTS_DIR / f"{today_str}_eod_report.xlsx"

    async with get_db_session() as session:
        pos_repo = PositionRepository(session)
        risk_repo = RiskRepository(session)

        positions = await pos_repo.get_all_active()
        var_result = await risk_repo.get_latest_var()

    wb = openpyxl.Workbook()

    # 포지션 시트
    ws_pos = wb.active
    ws_pos.title = "Positions"
    headers = ["ID", "기초자산", "구분", "C/P", "행사가", "만기", "수량", "평균가", "명목금액"]
    for col, header in enumerate(headers, 1):
        cell = ws_pos.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A1A2E")
        cell.alignment = Alignment(horizontal="center")

    for row, pos in enumerate(positions, 2):
        ws_pos.cell(row=row, column=1, value=pos.id)
        ws_pos.cell(row=row, column=2, value=pos.underlying)
        ws_pos.cell(row=row, column=3, value=pos.instrument_type)
        ws_pos.cell(row=row, column=4, value=pos.option_type)
        ws_pos.cell(row=row, column=5, value=pos.strike)
        ws_pos.cell(row=row, column=6, value=str(pos.expiry) if pos.expiry else "")
        ws_pos.cell(row=row, column=7, value=pos.quantity)
        ws_pos.cell(row=row, column=8, value=pos.avg_price)
        ws_pos.cell(row=row, column=9, value=pos.notional)

    # VaR 시트
    ws_var = wb.create_sheet("VaR")
    if var_result:
        ws_var["A1"] = "기준일"
        ws_var["B1"] = str(var_result.calc_date)
        ws_var["A2"] = "방법론"
        ws_var["B2"] = var_result.method
        ws_var["A3"] = "신뢰수준"
        ws_var["B3"] = f"{var_result.confidence:.0%}"
        ws_var["A4"] = "VaR"
        ws_var["B4"] = f"{var_result.var_amount:,.0f}원"
        ws_var["A5"] = "CVaR"
        ws_var["B5"] = f"{var_result.cvar_amount:,.0f}원"

    wb.save(output_path)
    logger.info("EOD 보고서 생성 완료: %s", output_path)
